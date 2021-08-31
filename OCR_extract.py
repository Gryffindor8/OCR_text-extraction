import fitz  # pymupdf
from os import walk
import os
import openpyxl
import pandas as pd
import tkinter as tk

cwd = os.getcwd()
root = tk.Tk()


def add_column(file_name, sheet_name, column):
    flee = pd.read_excel(file_name, engine="openpyxl")
    if "Area" not in flee.columns:
        wb = openpyxl.load_workbook(cwd + "//" + file_name)
        ws = wb[sheet_name]
        new_column = ws.max_column + 1
        for rowy, value1 in enumerate(column, start=1):
            ws.cell(row=rowy, column=new_column, value=value1)
        wb.save(file_name)
        lbl = tk.Label(root, text="Done", fg='red', font=("Helvetica", 10))
        lbl.place(x=60, y=50)
        lbl.pack()
    else:
        lbl = tk.Label(root, text="Exist already", fg='red', font=("Helvetica", 10))
        lbl.place(x=60, y=50)
        lbl.pack()


def main2():
    _, _, filename = next(walk(cwd))
    pdf = []
    excel = []
    # print(filename)
    if len(filename) >= 2:
        for k in filename:
            if k.endswith(".pdf"):
                pdf.append(k)
            if k.endswith(".xlsx"):
                excel.append(k)
        for j in pdf:
            for fle in excel:
                try:
                    if j.split()[0] in fle:
                        exc = pd.read_excel(fle, sheet_name="Sheet1", engine='openpyxl', header=0)
                        nme = (exc.columns[0])
                        rows = (len(exc[nme]))
                        area = ["Area"]
                        with (fitz.open(j)) as doc:
                            for page in doc:
                                text = page.getText().replace("\n", " ").strip()
                                if "SURFACE AREA" in text:
                                    try:
                                        tex = (text[text.index("SURFACE AREA"):text.index("sq.")])
                                        area.append(tex.replace("SURFACE AREA:", ""))
                                    except:
                                        area.append("Not Found")
                                else:
                                    area.append("Not Found")
                        try:
                            area = area[:rows + 1]
                        except:
                            pass
                        add_column(fle, "Sheet1", area)
                except IndexError:
                    print("Some Error occured")

    else:
        print("File Not Found")


def screen():
    root.geometry("200x100+10+20")
    frame = tk.Frame(root)
    frame.pack()
    button = tk.Button(frame, text="QUIT", fg="red", command=exit)
    button.pack(side=tk.RIGHT)
    slogan = tk.Button(frame, text="Start", command=main2)
    slogan.pack(side=tk.LEFT)
    root.mainloop()


screen()
